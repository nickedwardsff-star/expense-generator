import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import json
import base64
from openai import OpenAI
import fitz  
import re  

# --- 0. FARMFOODS BRANDING & PAGE CONFIG ---
st.set_page_config(page_title="Farmfoods Expenses", page_icon="🛒", layout="wide") 

st.markdown("""
    <style>
        .stApp { background-color: #f9fbf9; }
        
        p, label, .stMarkdown, .stText, .stCheckbox label, li { 
            color: #111111 !important; 
        }
        
        h1, h2, h3 { color: #007a33 !important; font-weight: 800 !important; }
        
        .stButton>button, .stDownloadButton>button {
            background-color: #007a33 !important;
            color: white !important;
            border-radius: 8px !important;
            border: none !important;
            font-weight: bold !important;
            transition: all 0.3s ease !important;
        }
        .stButton>button:hover, .stDownloadButton>button:hover {
            background-color: #da291c !important;
            color: white !important;
            transform: scale(1.02);
        }
        
        .stButton>button span, .stDownloadButton>button p {
            color: white !important;
        }
        
        .stFileUploader {
            border: 2px dashed #007a33 !important;
            border-radius: 10px !important;
            padding: 10px !important;
            background-color: #ffffff;
        }
        .stProgress > div > div > div > div {
            background-color: #007a33;
        }
    </style>
""", unsafe_allow_html=True)

# --- 1. SETUP & AUTHENTICATION ---
if 'expenses' not in st.session_state:
    st.session_state.expenses = []

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

# --- 2. THE REAL AI EXTRACTION ---
def to_float(val):
    try:
        if isinstance(val, str):
            val = val.replace('£', '').replace('$', '').replace(',', '').replace(' miles', '').strip()
        return round(float(val), 2)
    except:
        return 0.00

def safe_extract_json_from_response(response):
    try: 
        return response.choices.message.content
    except: 
        pass
        
    try: 
        return response['choices']['message']['content']
    except: 
        pass
        
    try: 
        return response.model_dump()['choices']['message']['content']
    except: 
        pass
    
    raw = str(response)
    match = re.search(r"content='(\{.*?\})', refusal", raw, re.DOTALL)
    if match: 
        return match.group(1).replace('\\n', '\n').replace('\\"', '"')
        
    match = re.search(r'content="(\{.*?\})", refusal', raw, re.DOTALL)
    if match: 
        return match.group(1).replace('\\n', '\n')
        
    return "{}"

def extract_receipt_data(uploaded_file, mileage_rate):
    extracted_text = ""
    base64_image = None
    
    if uploaded_file.type == 'application/pdf':
        doc = fitz.open(stream=uploaded_file.getvalue(), filetype="pdf")
        page = doc.load_page(0) 
        extracted_text = page.get_text().strip()
        
        if len(extracted_text) < 20:
            zoom_matrix = fitz.Matrix(1.5, 1.5) 
            pix = page.get_pixmap(matrix=zoom_matrix, alpha=False)
            base64_image = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
    else:
        base64_image = base64.b64encode(uploaded_file.getvalue()).decode('utf-8')

    prompt = (
        "You are an expert accountant. Analyze the uploaded document.\n"
        "If it is a SHOPPING RECEIPT, return a JSON object with: "
        "'Date' (YYYY-MM-DD), 'Vendor' (shop name), 'Reason' (2-4 words), "
        "'Amount Excl VAT' (number), 'VAT' (number), 'Total Amount' (number), 'Miles' (0).\n"
        "If it is a MAP or DRIVING DIRECTIONS for a mileage claim, return a JSON object with: "
        "'Date' (use today if missing), 'Vendor' (set to 'Mileage Claim'), "
        "'Reason' (Route summary, e.g., 'Glasgow to Home'), 'Amount Excl VAT' (0), 'VAT' (0), 'Total Amount' (0), "
        "'Miles' (extract the total distance in miles as a number)."
    )
    
    if len(extracted_text) >= 20:
        messages_payload = [
            {"type": "text", "text": prompt},
            {"type": "text", "text": "Document Text:\n" + extracted_text}
        ]
    else:
        messages_payload = [
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": "data:image/jpeg;base64," + base64_image, "detail": "high"}}
        ]
        
    response = client.chat.completions.create(
        model="gpt-4o",
        response_format={ "type": "json_object" },
        messages=[{"role": "user", "content": messages_payload}]
    )
    
    content = safe_extract_json_from_response(response)
        
    bticks = chr(96) * 3
    json_marker = bticks + "json"
    
    if json_marker in content: 
        content = content.split(json_marker).split(bticks).strip()
    elif bticks in content: 
        content = content.split(bticks).split(bticks).strip()
        
    try: 
        data = json.loads(content)
    except Exception: 
        data = {}
        
    while isinstance(data, list): 
        if len(data) > 0:
            data = data
        else:
            data = {}
        
    if isinstance(data, dict) and "Date" not in data:
        for key, value in data.items():
            if isinstance(value, dict) and "Date" in value:
                data = value
                break

    miles = to_float(data.get("Miles", 0))
    
    # THE FIX: Explicitly label the EV / Standard math in the Vendor column!
    if miles > 0 and mileage_rate > 0:
        calc_total = round(miles * mileage_rate, 2)
        rate_label = "EV @ 7p/mile" if mileage_rate == 0.07 else "Standard @ 30p/mile"
        
        clean_data = {
            "Date": str(data.get("Date") or datetime.now().strftime("%Y-%m-%d")),
            "Vendor": f"Mileage ({rate_label})",
            "Reason": str(data.get("Reason") or "Business Travel"),
            "Miles": miles,  
            "File Name": uploaded_file.name,
            "Amount Excl VAT": calc_total,
            "VAT": 0.00,
            "Total Amount": calc_total
        }
    else:
        clean_data = {
            "Date": str(data.get("Date") or datetime.now().strftime("%Y-%m-%d")),
            "Vendor": str(data.get("Vendor") or "Unknown Vendor"),
            "Reason": str(data.get("Reason") or "General Expense"),
            "Miles": 0.00,   
            "File Name": uploaded_file.name,
            "Amount Excl VAT": to_float(data.get("Amount Excl VAT")),
            "VAT": to_float(data.get("VAT")),
            "Total Amount": to_float(data.get("Total Amount"))
        }
    
    return clean_data

# --- 3. USER INTERFACE ---
st.title("🛒 Farmfoods Expense Generator")

st.write("Upload your receipts and maps to automatically extract data, assign audit references, and build your submission pack.")

st.info("👋 **Welcome!** Please enter your full name below, and then upload all of your receipt files for the month to build your audited submission pack.")

col1, col2 = st.columns(2)
with col1:
    st.write("**Employee Details**")
    st.caption("Please provide the name for the final report.")
    employee_name = st.text_input("Enter your full name:", placeholder="e.g., Jane Doe", label_visibility="collapsed")
    
with col2:
    st.write("**Mileage Claims**")
    st.caption("Driving for work? Check the box below to automatically calculate your maps.")
    claiming_mileage = st.checkbox("🚗 Yes, I am claiming mileage this month")
    
    if claiming_mileage:
        car_selection = st.selectbox("Select your vehicle type:", ["Hybrid / Petrol / Diesel (30p / mile)", "Electric Vehicle (7p / mile)"])
        mileage_rate = 0.07 if "Electric" in car_selection else 0.30
    else:
        mileage_rate = 0.00

st.divider()

uploaded_files = st.file_uploader("Upload Receipts & Maps Screenshots", type=['png', 'jpg', 'jpeg', 'pdf'], accept_multiple_files=True)

if uploaded_files and employee_name:
    if st.button("Process " + str(len(uploaded_files)) + " File(s)"):
        
        progress_text = st.empty()
        progress_bar = st.progress(0)
        total_files = len(uploaded_files)
        
        for i, file in enumerate(uploaded_files):
            progress_text.text("Reading file " + str(i + 1) + " of " + str(total_files) + "...")
            try:
                extracted_data = extract_receipt_data(file, mileage_rate)
                st.session_state.expenses.append(extracted_data)
            except Exception as e:
                st.warning(f"Could not extract data from {file.name}: {str(e)}")
            progress_bar.progress((i + 1) / total_files)
            
        progress_text.text("Finished reading! Building your audited files...")
        st.success("All done!")

# --- 4. DISPLAY AND TEMPLATE DOWNLOAD ---
if len(st.session_state.expenses) > 0:
    st.divider()
    st.subheader("Current Report for " + employee_name)
    
    df = pd.DataFrame(st.session_state.expenses)
    df = df.sort_values(by="Date").reset_index(drop=True)
    df["Reference"] = [str(i) for i in range(1, len(df) + 1)] 
    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d/%m/%Y")
    
    display_cols = ["Reference", "Date", "Vendor", "Reason", "Miles", "Amount Excl VAT", "VAT", "Total Amount"]
    st.dataframe(df[display_cols], use_container_width=True)
    
    try:
        # 2. Build the Excel File 
        wb = openpyxl.load_workbook("Template_Expenses.xlsx") 
        ws = wb.active 
        
        ws['B3'] = employee_name 
        ws['H3'] = datetime.now().strftime("%d/%m/%Y")
        
        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        center_align = Alignment(horizontal='center')
        
        start_row = 6 
        
        for index, row in df.iterrows():
            current_row = start_row + index
            audited_vendor = "REF " + str(row["Reference"]) + " - " + str(row["Vendor"])
            
            cell_date = ws.cell(row=current_row, column=1, value=row["Date"])
            cell_date.alignment = left_align
            
            cell_vendor = ws.cell(row=current_row, column=2, value=audited_vendor)
            cell_vendor.alignment = left_align
            
            cell_reason = ws.cell(row=current_row, column=3, value=row["Reason"])
            cell_reason.alignment = left_align
            
            if row["Miles"] > 0:
                cell_miles = ws.cell(row=current_row, column=4, value=row["Miles"])
                cell_miles.alignment = center_align
            
            cell_excl = ws.cell(row=current_row, column=5, value=row["Amount Excl VAT"])
            cell_excl.alignment = right_align
            cell_excl.number_format = '#,##0.00'
            
            cell_vat = ws.cell(row=current_row, column=6, value=row["VAT"])
            cell_vat.alignment = right_align
            cell_vat.number_format = '#,##0.00'
            
            cell_total = ws.cell(row=current_row, column=7, value=row["Total Amount"])
            cell_total.alignment = right_align
            cell_total.number_format = '#,##0.00'
            
        totals_row = start_row + len(df) + 1 
        bold_font = Font(bold=True)
        
        ws.cell(row=totals_row, column=4, value="GRAND TOTAL").font = bold_font
        ws.cell(row=totals_row, column=4).alignment = right_align
        
        total_excl = ws.cell(row=totals_row, column=5, value=df["Amount Excl VAT"].sum())
        total_excl.font = bold_font
        total_excl.alignment = right_align
        total_excl.number_format = '#,##0.00'
        
        total_vat = ws.cell(row=totals_row, column=6, value=df["VAT"].sum())
        total_vat.font = bold_font
        total_vat.alignment = right_align
        total_vat.number_format = '#,##0.00'
        
        total_final = ws.cell(row=totals_row, column=7, value=df["Total Amount"].sum())
        total_final.font = bold_font
        total_final.alignment = right_align
        total_final.number_format = '#,##0.00'
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        
        # 3. Build the Master PDF pack (THE BULLETPROOF HYBRID METHOD)
        master_pdf = fitz.open() 
        
        for index, row in df.iterrows():
            filename = row["File Name"]
            ref_id = row["Reference"]
            
            file_obj = next((f for f in uploaded_files if f.name == filename), None)
            
            if file_obj:
                ext = file_obj.name.split('.')[-1].lower()
                try:
                    if ext == 'pdf':
                        source_doc = fitz.open(stream=file_obj.getvalue(), filetype="pdf")
                    elif ext in ['jpg', 'jpeg', 'png']:
                        source_doc = fitz.open(stream=file_obj.getvalue(), filetype=ext)
                    else:
                        continue
                    
                    for page_num in range(len(source_doc)):
                        page = source_doc.load_page(page_num)
                        
                        zoom_matrix = fitz.Matrix(1.5, 1.5)
                        pix = page.get_pixmap(matrix=zoom_matrix, alpha=False)
                        
                        img_bytes = pix.tobytes("jpeg")
                        
                        img_doc = fitz.open(stream=img_bytes, filetype="jpeg")
                        pdf_bytes_from_img = img_doc.convert_to_pdf()
                        
                        flat_doc = fitz.open("pdf", pdf_bytes_from_img)
                        flat_page = flat_doc.load_page(0) 
                        
                        if page_num == 0:
                            try:
                                bg_rect = fitz.Rect(10, 10, 250, 70)
                                if hasattr(flat_page, "draw_rect"):
                                    flat_page.draw_rect(bg_rect, color=(1, 1, 1), fill=(1, 1, 1))
                                elif hasattr(flat_page, "drawRect"):
                                    flat_page.drawRect(bg_rect, color=(1, 1, 1), fill=(1, 1, 1))
                                
                                target_point = fitz.Point(20, 50)
                                ref_text = "REF: " + str(ref_id)
                                red_color = (0.85, 0.16, 0.11)
                                
                                if hasattr(flat_page, "insert_text"):
                                    flat_page.insert_text(target_point, ref_text, fontsize=30, color=red_color)
                                elif hasattr(flat_page, "insertText"):
                                    flat_page.insertText(target_point, ref_text, fontsize=30, color=red_color)
                                    
                            except Exception as e:
                                st.warning(f"Could not stamp {filename}: {str(e)}")
                                
                        master_pdf.insert_pdf(flat_doc)
                        
                        img_doc.close()
                        flat_doc.close()
                                
                except Exception as e:
                    st.warning(f"Failed to process {filename}: {str(e)}")
                    
        # 4. Create the dual download buttons
        safe_name = employee_name.replace(" ", "_")
        
        st.write("### Download Your Audited Files")
        col1, col2 = st.columns(2)
        
        with col1:
            st.download_button(
                label="📊 Download Excel Spreadsheet",
                data=excel_buffer.getvalue(),
                file_name="Farmfoods_Expense_Report_" + safe_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
        with col2:
            if master_pdf.page_count > 0:
                pdf_bytes = master_pdf.tobytes()
                st.download_button(
                    label="📑 Download Audited Receipt Pack",
                    data=pdf_bytes,
                    file_name="Farmfoods_Audited_Receipts_" + safe_name + ".pdf",
                    mime="application/pdf",
                    type="primary",
                    use_container_width=True
                )
        
    except FileNotFoundError:
        st.error("⚠️ Could not find 'Template_Expenses.xlsx'. Please make sure it is saved in the same folder as this script.")
    
    st.divider()
    if st.button("Clear Data and Start Over"):
        st.session_state.expenses = []
        st.rerun()
